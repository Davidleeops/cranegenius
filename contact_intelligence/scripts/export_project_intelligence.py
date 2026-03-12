#!/usr/bin/env python3
"""
Export Project Intelligence candidate views for campaign usage and static frontend.
"""

from __future__ import annotations

import argparse
import json
import os
import sqlite3
from pathlib import Path

DEFAULT_DB = os.path.expanduser("~/data_runtime/cranegenius_ci.db")
DEFAULT_OUT = os.path.expanduser("~/data_runtime/exports")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    import csv as _csv


BASE_COLUMNS = """
    project_name_raw AS project_name,
    project_type,
    vertical,
    city,
    state,
    company_name_raw AS company_name,
    source_count,
    signal_count,
    earliest_signal_date,
    latest_signal_date,
    estimated_spend_proxy,
    crane_relevance_score,
    mini_crane_fit_score,
    confidence_score,
    monetization_score,
    recommendation_reason,
    priority_reason
"""

QUERIES = {
    "national_opportunity_candidates": f"""
        SELECT {BASE_COLUMNS}
        FROM project_candidates
        WHERE status='active'
        ORDER BY monetization_score DESC, confidence_score DESC
    """,
    "top_project_candidates": f"""
        SELECT {BASE_COLUMNS}
        FROM project_candidates
        WHERE status='active'
        ORDER BY monetization_score DESC, confidence_score DESC
        LIMIT 20
    """,
    "top_data_center_candidates": f"""
        SELECT {BASE_COLUMNS}
        FROM project_candidates
        WHERE vertical='data_centers' AND status='active'
        ORDER BY monetization_score DESC, confidence_score DESC
        LIMIT 50
    """,
    "top_energy_candidates": f"""
        SELECT {BASE_COLUMNS}
        FROM project_candidates
        WHERE vertical='power_energy' AND status='active'
        ORDER BY monetization_score DESC, confidence_score DESC
        LIMIT 50
    """,
    "top_industrial_candidates": f"""
        SELECT {BASE_COLUMNS}
        FROM project_candidates
        WHERE vertical='industrial_manufacturing' AND status='active'
        ORDER BY monetization_score DESC, confidence_score DESC
        LIMIT 50
    """,
    "top_logistics_candidates": f"""
        SELECT {BASE_COLUMNS}
        FROM project_candidates
        WHERE vertical='warehousing_logistics' AND status='active'
        ORDER BY monetization_score DESC, confidence_score DESC
        LIMIT 50
    """,
    "mini_opportunity_candidates": f"""
        SELECT {BASE_COLUMNS}
        FROM project_candidates
        WHERE mini_crane_fit_score >= 62 AND status='active'
        ORDER BY mini_crane_fit_score DESC, confidence_score DESC
        LIMIT 100
    """,
    "recommended_expansion_candidates": f"""
        SELECT {BASE_COLUMNS}
        FROM project_candidates
        WHERE (recommended_flag = 1 OR recommendation_reason <> '') AND status='active'
        ORDER BY monetization_score DESC, confidence_score DESC
        LIMIT 100
    """,
}


def write_xlsx(rows, headers, sheet_name, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="C9A84C")
    hfont = Font(bold=True, color="080E1A")
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 48)
    wb.save(str(out_path))


def write_csv(rows, headers, out_path):
    with open(out_path, "w", newline="") as f:
        w = _csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def run_export(conn: sqlite3.Connection, name: str, sql: str, out_dir: Path) -> int:
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    headers = [d[0] for d in cur.description]
    ext = "xlsx" if HAS_XLSX else "csv"
    out_path = out_dir / f"{name}.{ext}"
    if HAS_XLSX:
        write_xlsx(rows, headers, name, out_path)
    else:
        write_csv(rows, headers, out_path)
    print(f"[project-export] {name}.{ext}: {len(rows)} rows -> {out_path}")
    return len(rows)


def run_static_json(conn: sqlite3.Connection, out_dir: Path) -> None:
    static_dir = out_dir / "static_exports"
    static_dir.mkdir(parents=True, exist_ok=True)

    mapping = {
        "project_candidates.json": QUERIES["national_opportunity_candidates"],
        "top_project_candidates.json": QUERIES["top_project_candidates"],
        "mini_opportunity_candidates.json": QUERIES["mini_opportunity_candidates"],
        "recommended_expansion_candidates.json": QUERIES["recommended_expansion_candidates"],
    }

    cur = conn.cursor()
    for filename, sql in mapping.items():
        cur.execute(sql)
        headers = [d[0] for d in cur.description]
        rows = [dict(zip(headers, r)) for r in cur.fetchall()]
        path = static_dir / filename
        path.write_text(json.dumps(rows, indent=2), encoding="utf-8")
        print(f"[project-export] {filename}: {len(rows)} records -> {path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export Project Intelligence views and static JSON payloads.")
    parser.add_argument("--db", default=None)
    parser.add_argument("--output", default=None)
    parser.add_argument("--json-output", default=None, help="Directory for static JSON exports")
    args = parser.parse_args()

    db_path = args.db or os.environ.get("CRANEGENIUS_CI_DB", DEFAULT_DB)
    out_dir = Path(args.output or DEFAULT_OUT)
    json_out = Path(args.json_output) if args.json_output else Path(__file__).resolve().parents[2] / "data"

    if not Path(db_path).exists():
        raise SystemExit(f"DB not found: {db_path}")

    out_dir.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)

    counts = {}
    for name, sql in QUERIES.items():
        counts[name] = run_export(conn, name, sql, out_dir)

    run_static_json(conn, json_out)
    conn.close()

    print("[project-export] done")
    print("[project-export] counts:", counts)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
