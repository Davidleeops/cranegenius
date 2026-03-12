#!/usr/bin/env python3
"""
contact_intelligence/scripts/export_views.py
Generate .xlsx (or csv fallback) exports from the contact intelligence database.
"""

import os
import sys
import sqlite3
import argparse
from pathlib import Path

DEFAULT_DB = os.path.expanduser("~/data_runtime/cranegenius_ci.db")
DEFAULT_OUT = os.path.expanduser("~/data_runtime/exports")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    import csv as _csv

GOLD = "C9A84C"
DARK = "080E1A"

QUERIES = {
    "top_100_targets": """
        SELECT
            s.sector_name AS sector,
            co.company_name AS company,
            co.domain,
            co.company_type,
            co.location_city AS city,
            co.location_state AS state,
            co.region,
            co.target_tier AS tier,
            c.full_name AS contact_name,
            c.title,
            c.contact_role AS role,
            c.email,
            c.email_verified AS verified,
            c.phone,
            c.linkedin_url AS linkedin,
            c.confidence_score AS confidence
        FROM contacts c
        JOIN companies co ON co.company_id = c.company_id
        LEFT JOIN sectors s ON s.sector_id = co.sector_id
        WHERE c.email IS NOT NULL
        ORDER BY c.confidence_score DESC, c.email_verified DESC
        LIMIT 100
    """,
    "verified_contacts": """
        SELECT
            co.company_name AS company,
            co.domain,
            co.industry,
            s.sector_name AS sector,
            co.location_city AS city,
            co.location_state AS state,
            c.full_name AS contact_name,
            c.title,
            c.contact_role AS role,
            c.email,
            c.phone,
            c.linkedin_url AS linkedin,
            c.confidence_score AS confidence,
            c.last_verified_at
        FROM contacts c
        JOIN companies co ON co.company_id = c.company_id
        LEFT JOIN sectors s ON s.sector_id = co.sector_id
        WHERE c.email_verified = 1
        ORDER BY c.confidence_score DESC
    """,
    "shared_contacts": """
        SELECT
            c.full_name AS contact_name,
            c.title,
            co.company_name AS company,
            co.domain,
            c.email,
            GROUP_CONCAT(DISTINCT oh.business_line) AS business_lines,
            COUNT(DISTINCT oh.business_line) AS line_count
        FROM contacts c
        JOIN outreach_history oh ON oh.contact_id = c.contact_id
        LEFT JOIN companies co ON co.company_id = c.company_id
        GROUP BY c.contact_id
        HAVING COUNT(DISTINCT oh.business_line) > 1
        ORDER BY line_count DESC
    """,
    "high_confidence_targets": """
        SELECT
            s.sector_name AS sector,
            co.company_name AS company,
            co.domain,
            co.location_state AS state,
            co.region,
            c.full_name AS contact_name,
            c.title,
            c.email,
            c.email_verified AS verified,
            c.phone,
            c.confidence_score AS confidence,
            oh.campaign_name,
            oh.response_status
        FROM contacts c
        JOIN companies co ON co.company_id = c.company_id
        LEFT JOIN sectors s ON s.sector_id = co.sector_id
        LEFT JOIN outreach_history oh ON oh.contact_id = c.contact_id
        WHERE c.confidence_score >= 0.7 AND c.email IS NOT NULL
        ORDER BY c.confidence_score DESC
    """,
    "data_center_targets": """
        SELECT
            co.company_name AS company,
            co.domain,
            co.location_city AS city,
            co.location_state AS state,
            c.full_name AS contact_name,
            c.title,
            c.email,
            c.email_verified AS verified,
            c.phone,
            p.project_name,
            p.project_stage,
            p.estimated_project_value AS project_value,
            cr.crane_type_needed,
            cr.capacity_required
        FROM companies co
        LEFT JOIN sectors s ON s.sector_id = co.sector_id
        LEFT JOIN contacts c ON c.company_id = co.company_id
        LEFT JOIN projects p ON p.company_id = co.company_id
        LEFT JOIN crane_requirements cr ON cr.project_id = p.project_id
        WHERE s.sector_name LIKE '%Data Center%'
           OR co.industry LIKE '%data center%'
           OR p.project_type = 'data_center'
        ORDER BY p.crane_need_score DESC, c.confidence_score DESC
    """,
    "industrial_turnaround": """
        SELECT
            co.company_name AS company,
            co.domain,
            co.location_city AS city,
            co.location_state AS state,
            c.full_name AS contact_name,
            c.title,
            c.email,
            c.phone,
            sig.signal_type,
            sig.signal_value,
            sig.signal_date,
            sig.signal_confidence
        FROM signals sig
        JOIN companies co ON co.company_id = sig.company_id
        LEFT JOIN contacts c ON c.company_id = co.company_id
        WHERE sig.signal_type IN ('industrial_turnaround','shutdown','turnaround')
           OR co.industry LIKE '%industrial%'
           OR co.industry LIKE '%refinery%'
        ORDER BY sig.signal_confidence DESC, sig.captured_at DESC
    """,
    "usable_contacts_all_sources": """
        WITH latest_facts AS (
            SELECT f.*
            FROM contact_source_facts f
            JOIN (
                SELECT contact_id, MAX(last_seen_at) AS max_seen
                FROM contact_source_facts
                GROUP BY contact_id
            ) m ON m.contact_id = f.contact_id AND m.max_seen = f.last_seen_at
        )
        SELECT
            lf.source_system,
            lf.source_file,
            lf.source_sheet,
            lf.source_row,
            c.contact_id,
            c.full_name AS contact_name,
            c.title,
            c.email,
            co.company_name,
            COALESCE(csf.preferred_domain, co.domain) AS preferred_domain,
            lf.email_verification_status,
            lf.email_domain_type,
            lf.title_confidence,
            lf.person_confidence,
            lf.company_confidence,
            lf.record_quality_score,
            lf.usable_reason,
            lf.notes,
            lf.last_seen_at
        FROM latest_facts lf
        JOIN contacts c ON c.contact_id = lf.contact_id
        LEFT JOIN companies co ON co.company_id = c.company_id
        LEFT JOIN company_source_facts csf ON csf.company_id = co.company_id AND csf.source_system = lf.source_system
        WHERE lf.usable_for_outreach = 1
        ORDER BY lf.record_quality_score DESC, lf.last_seen_at DESC
    """,
    "unusable_contacts_with_reasons": """
        WITH latest_facts AS (
            SELECT f.*
            FROM contact_source_facts f
            JOIN (
                SELECT contact_id, MAX(last_seen_at) AS max_seen
                FROM contact_source_facts
                GROUP BY contact_id
            ) m ON m.contact_id = f.contact_id AND m.max_seen = f.last_seen_at
        )
        SELECT
            lf.source_system,
            lf.source_file,
            lf.source_sheet,
            lf.source_row,
            c.contact_id,
            c.full_name AS contact_name,
            c.email,
            co.company_name,
            lf.email_verification_status,
            lf.email_domain_type,
            lf.record_quality_score,
            lf.blocked_reason,
            lf.notes,
            lf.last_seen_at
        FROM latest_facts lf
        JOIN contacts c ON c.contact_id = lf.contact_id
        LEFT JOIN companies co ON co.company_id = c.company_id
        WHERE lf.usable_for_outreach = 0
        ORDER BY lf.last_seen_at DESC
    """,
    "company_domain_seed_candidates": """
        SELECT
            co.company_id,
            co.company_name,
            co.normalized_company_name,
            csf.source_system,
            csf.preferred_domain,
            csf.source_support_count,
            csf.domain_confidence,
            csf.domain_evidence_notes,
            csf.quality_notes,
            csf.last_seen_at
        FROM company_source_facts csf
        JOIN companies co ON co.company_id = csf.company_id
        WHERE COALESCE(csf.preferred_domain, '') <> ''
          AND csf.domain_confidence >= 0.55
        ORDER BY csf.domain_confidence DESC, csf.source_support_count DESC, csf.last_seen_at DESC
    """,
    "source_quality_summary": """
        WITH base AS (
            SELECT
                source_system,
                contact_id,
                usable_for_outreach,
                email_verification_status,
                email_domain_type,
                blocked_reason
            FROM contact_source_facts
        ),
        blocked AS (
            SELECT
                source_system,
                blocked_reason,
                COUNT(*) AS cnt
            FROM base
            WHERE COALESCE(blocked_reason, '') <> ''
            GROUP BY source_system, blocked_reason
        ),
        blocked_rollup AS (
            SELECT
                source_system,
                GROUP_CONCAT(blocked_reason || ':' || cnt, '; ') AS blocked_reason_breakdown
            FROM blocked
            GROUP BY source_system
        )
        SELECT
            b.source_system,
            COUNT(DISTINCT b.contact_id) AS total_contacts,
            COUNT(DISTINCT CASE WHEN b.usable_for_outreach = 1 THEN b.contact_id END) AS usable_contacts,
            ROUND(
                100.0 * COUNT(CASE WHEN b.email_verification_status IN ('valid','catchall') THEN 1 END)
                / NULLIF(COUNT(*), 0),
                2
            ) AS verification_rate_pct,
            COUNT(CASE WHEN b.email_domain_type = 'corporate' THEN 1 END) AS corporate_email_rows,
            COUNT(CASE WHEN b.email_domain_type = 'free_or_isp' THEN 1 END) AS free_email_rows,
            COALESCE(br.blocked_reason_breakdown, '') AS blocked_reason_breakdown
        FROM base b
        LEFT JOIN blocked_rollup br ON br.source_system = b.source_system
        GROUP BY b.source_system, br.blocked_reason_breakdown
        ORDER BY usable_contacts DESC, total_contacts DESC
    """,

    "job_contact_match_candidates": """
        SELECT
            j.job_feed_id,
            j.title AS job_title,
            j.company_name AS job_company,
            j.location_state AS job_state,
            m.match_score,
            m.match_reason,
            c.contact_id,
            c.full_name AS contact_name,
            c.title AS contact_title,
            c.email,
            co.company_name AS contact_company
        FROM job_contact_matches m
        JOIN jobs_feed_items j ON j.job_feed_id = m.job_feed_id
        JOIN contacts c ON c.contact_id = m.contact_id
        LEFT JOIN companies co ON co.company_id = c.company_id
        ORDER BY m.match_score DESC, j.job_feed_id ASC
    """,
    "opportunity_company_match_candidates": """
        SELECT
            o.opportunity_feed_id,
            o.project_name,
            o.city,
            o.location_state,
            o.opportunity_type,
            m.match_score,
            m.match_reason,
            co.company_id,
            co.company_name,
            co.domain,
            co.company_type,
            co.industry
        FROM opportunity_company_matches m
        JOIN opportunity_feed_items o ON o.opportunity_feed_id = m.opportunity_feed_id
        JOIN companies co ON co.company_id = m.company_id
        ORDER BY m.match_score DESC, o.opportunity_feed_id ASC
    """,
    "manpower_job_match_candidates": """
        SELECT
            p.manpower_profile_id,
            p.full_name,
            p.title AS profile_title,
            p.location_state AS profile_state,
            j.job_feed_id,
            j.title AS job_title,
            j.company_name,
            j.location_state AS job_state,
            m.match_score,
            m.match_reason
        FROM manpower_job_matches m
        JOIN manpower_profiles p ON p.manpower_profile_id = m.manpower_profile_id
        JOIN jobs_feed_items j ON j.job_feed_id = m.job_feed_id
        ORDER BY m.match_score DESC, p.manpower_profile_id ASC
    """,

}


def write_xlsx(rows, headers, sheet_name, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor=GOLD)
    hfont = Font(bold=True, color=DARK)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col in ws.columns:
        w = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 50)
    wb.save(str(out_path))


def write_csv_fallback(rows, headers, out_path):
    with open(out_path, "w", newline="") as f:
        w = _csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def run_export(conn, name, sql, out_dir):
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    headers = [d[0] for d in cur.description]
    ext = "xlsx" if HAS_XLSX else "csv"
    out_path = Path(out_dir) / f"{name}.{ext}"
    if HAS_XLSX:
        write_xlsx(rows, headers, name[:31], out_path)
    else:
        write_csv_fallback(rows, headers, out_path)
    print(f"  [export] {name}.{ext} — {len(rows)} rows → {out_path}")


def run_sector_export(conn, out_dir):
    cur = conn.cursor()
    cur.execute("SELECT sector_id, sector_name FROM sectors WHERE active=1 ORDER BY display_order")
    sectors = cur.fetchall()
    if not sectors:
        return
    if not HAS_XLSX:
        print("  [export] top100_by_sector skipped (openpyxl required)")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hfill = PatternFill("solid", fgColor=GOLD)
    hfont = Font(bold=True, color=DARK)

    for sid, sname in sectors:
        cur.execute(
            """
            SELECT co.company_name, co.domain, co.location_city, co.location_state,
                   co.company_type, co.target_tier, co.target_score,
                   c.full_name, c.title, c.email, c.email_verified, c.phone
            FROM companies co
            LEFT JOIN contacts c ON c.company_id = co.company_id
            WHERE co.sector_id=?
            ORDER BY co.target_score DESC, c.confidence_score DESC
            LIMIT 100
            """,
            (sid,),
        )
        rows = cur.fetchall()
        if not rows:
            continue
        ws = wb.create_sheet(title=sname[:31])
        headers = ["company", "domain", "city", "state", "type", "tier", "score", "contact", "title", "email", "verified", "phone"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(list(row))
        for col in ws.columns:
            w = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)

    if wb.sheetnames:
        out_path = Path(out_dir) / "top100_by_sector.xlsx"
        wb.save(str(out_path))
        print(f"  [export] top100_by_sector.xlsx — {len(wb.sheetnames)} sector tabs")


def main(db_path, out_dir, single=None):
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)

    to_run = {single: QUERIES[single]} if single and single in QUERIES else QUERIES
    if single and single not in QUERIES and single != "top100_by_sector":
        sys.exit(f"[export] Unknown export: {single}. Options: {list(QUERIES.keys()) + ['top100_by_sector']}")

    print(f"[export] Generating exports to {out_dir}...")
    for name, sql in to_run.items():
        try:
            run_export(conn, name, sql, out_dir)
        except Exception as e:
            print(f"  [export] ERROR {name}: {e}")

    if not single or single == "top100_by_sector":
        try:
            run_sector_export(conn, out_dir)
        except Exception as e:
            print(f"  [export] ERROR top100_by_sector: {e}")

    conn.close()
    print("[export] ✓ Done.")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--db", default=None)
    p.add_argument("--output", default=None)
    p.add_argument("--export", default=None)
    a = p.parse_args()

    db = a.db or os.environ.get("CRANEGENIUS_CI_DB", DEFAULT_DB)
    out = a.output or DEFAULT_OUT
    if not Path(db).exists():
        sys.exit(f"[export] DB not found: {db}")
    main(db, out, a.export)
