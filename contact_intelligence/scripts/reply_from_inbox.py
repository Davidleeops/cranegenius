#!/usr/bin/env python3
from __future__ import annotations
import csv, re, argparse
from pathlib import Path
from openpyxl import load_workbook

def n(v): return str(v or "").strip()
def l(v): return n(v).lower()

TEMPLATES = {
    "hot_spider": """Appreciate the reply.

Just so I send the right thing back, are you actually running spider crane work on this one, and what city is the job in

Also is this happening this week or more this month""",
    "hot_broad": """Appreciate it.

Just so I point you the right direction, is this a tight access lift situation or more of a general equipment need

What city is the job in, and is the timing this week or later this month""",
    "warm": """Makes sense.

Do these tight access or interior lift jobs come up for your team at all

If so, what city are you mainly operating in right now""",
    "related_not_spider": """Good to know.

Even if it is not spider specific, do you run into jobs where setup room is tight and a normal lift plan does not work

If yes, what type of projects usually trigger that""",
    "wrong_contact": """Thanks for letting me know.

Who normally handles crane or lift coordination on your side

I will send it directly to them and keep it short""",
    "not_fit": """Got it, appreciate the quick response.

If something tight access comes up later feel free to send over the city and timing""",
    "do_not_contact": "",
}

def classify(text: str) -> str:
    t = l(text)
    if any(x in t for x in ["unsubscribe", "remove me", "do not contact", "stop emailing", "spam"]):
        return "do_not_contact"
    if any(x in t for x in ["wrong person", "not the right person", "no longer with", "left the company"]):
        return "wrong_contact"
    spider = any(x in t for x in ["spider crane", "urw295", "unic"])
    positive = any(x in t for x in ["yes", "interested", "can help", "we do", "we handle", "available", "send details", "tell me more"])
    negative = any(x in t for x in ["not interested", "no thanks", "we don't", "do not", "nope"])
    related = any(x in t for x in ["glazing", "glass", "rigging", "interior", "tight access", "limited access", "facade"])
    if negative and not related:
        return "not_fit"
    if spider and positive:
        return "hot_spider"
    if positive:
        return "hot_broad"
    if related:
        return "related_not_spider"
    return "warm"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", default="runs/reply_drafts_from_inbox.csv")
    args = ap.parse_args()

    xlsx = Path(args.xlsx).expanduser().resolve()
    out = Path(args.out).resolve()

    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("No rows found")

    headers = [n(h) for h in rows[0]]
    data = rows[1:]

    text_cols = [i for i,h in enumerate(headers) if any(k in l(h) for k in ["reply", "message", "body", "text", "content"])]
    sender_col = next((i for i,h in enumerate(headers) if any(k in l(h) for k in ["from", "sender", "email"])), None)
    subject_col = next((i for i,h in enumerate(headers) if "subject" in l(h)), None)

    if not text_cols:
        text_cols = list(range(min(3, len(headers))))

    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "source_row", "sender", "subject", "reply_bucket", "suggested_reply",
                "next_question_1", "next_question_2", "next_question_3", "next_action"
            ]
        )
        w.writeheader()

        for idx, r in enumerate(data, start=2):
            text = " ".join(n(r[c]) for c in text_cols if c < len(r))
            bucket = classify(text)
            sender = n(r[sender_col]) if sender_col is not None and sender_col < len(r) else ""
            subject = n(r[subject_col]) if subject_col is not None and subject_col < len(r) else ""

            q1 = "Do you handle this type of lift work"
            q2 = "What city is the job in"
            q3 = "Is timing this week, this month, or later"
            action = {
                "hot_spider": "Send URW295 details and confirm timing",
                "hot_broad": "Identify access constraint and equipment need",
                "warm": "Capture city and timing, keep thread active",
                "related_not_spider": "Clarify lift type and project trigger",
                "wrong_contact": "Request correct contact and update record",
                "not_fit": "Close politely and keep door open",
                "do_not_contact": "Suppress contact immediately",
            }[bucket]

            w.writerow({
                "source_row": idx,
                "sender": sender,
                "subject": subject,
                "reply_bucket": bucket,
                "suggested_reply": TEMPLATES[bucket],
                "next_question_1": q1 if bucket not in {"wrong_contact","not_fit","do_not_contact"} else "",
                "next_question_2": q2 if bucket not in {"wrong_contact","not_fit","do_not_contact"} else "",
                "next_question_3": q3 if bucket not in {"wrong_contact","not_fit","do_not_contact"} else "",
                "next_action": action,
            })

    print(f"Wrote: {out}")

if __name__ == "__main__":
    main()
